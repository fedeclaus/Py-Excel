import xlsxwriter
from openpyxl import Workbook,load_workbook

from openpyxl.workbook import Workbook
from openpyxl.utils.cell import range_boundaries

def unm_wb(wb):
    for st_name in wb.sheetnames:
        st = wb[st_name]
        mcr_coord_list = [mcr.coord for mcr in st.merged_cells.ranges]

        for mcr in mcr_coord_list:
            min_col, min_row, max_col, max_row = range_boundaries(mcr)
            top_left_cell_value = st.cell(row=min_row, column=min_col).value
            if min_col <= 3 and max_col >= 9:
                st.unmerge_cells(mcr)
                for row in st.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                    for cell in row:
                        cell.value = top_left_cell_value
    return wb

input = load_workbook(filename='input.xlsx')
data = input['Sheet1']
output = load_workbook(filename='template.xlsx')
#data2= output['Charter']
data2 = unm_wb(output)
data2 = data2['Charter']

#print(data['B2'].value)

#data2.unmerge_cells(start_row=1, start_column=3, end_row=1, end_column=4)
data2.cell(1,3).value = data['A2'].value
data2.cell(2,3).value = data['B2'].value
data2.cell(3,3).value = data['C2'].value
data2.cell(6,3).value = data['D2'].value
data2.merge_cells('C6:I9')
data2.cell(11,3).value = data['E2'].value
data2.merge_cells('C11:I14')
data2.cell(16,3).value = data['F2'].value
data2.merge_cells('C16:I19')
data2.cell(21,3).value = data['G2'].value
data2.merge_cells('C21:I24')
data2.cell(26,5).value = data['H2'].value
data2.merge_cells('E26:I26')
data2.cell(27,5).value = data['I2'].value
data2.merge_cells('E27:I27')
data2.cell(28,5).value = data['J2'].value
data2.merge_cells('E28:I28')
data2.cell(29,5).value = data['K2'].value
data2.merge_cells('E29:I29')

data2.cell(32,3).value = data['L2'].value
data2.cell(33,3).value = data['M2'].value
data2.cell(34,3).value = data['N2'].value
data2.cell(32,7).value = data['O2'].value
data2.cell(33,7).value = data['P2'].value
data2.cell(34,7).value = data['Q2'].value

data2.cell(37,5).value = data['R2'].value
data2.cell(38,5).value = data['S2'].value
data2.cell(39,5).value = data['T2'].value
data2.cell(41,3).value = data['U2'].value
data2.merge_cells('C41:I53')

output.save(filename='template2.xlsx')
